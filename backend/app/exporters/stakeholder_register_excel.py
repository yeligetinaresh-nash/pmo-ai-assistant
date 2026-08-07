from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _list_to_text(values: list[str]) -> str:
    if not values:
        return ""

    return "\n".join(
        f"• {value}"
        for value in values
    )


def _apply_header_style(
    worksheet,
    row_number: int,
) -> None:
    for cell in worksheet[row_number]:
        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )


def _set_column_widths(
    worksheet,
    widths: dict[str, float],
) -> None:
    for column_letter, width in widths.items():
        worksheet.column_dimensions[
            column_letter
        ].width = width


def generate_stakeholder_register_excel(
    content: dict,
) -> BytesIO:
    workbook = Workbook()

    register_sheet = workbook.active
    register_sheet.title = "Stakeholder Register"

    summary_sheet = workbook.create_sheet(
        "Summary"
    )

    assumptions_sheet = workbook.create_sheet(
        "Assumptions and Notes"
    )

    register_headers = [
        "Stakeholder ID",
        "Name",
        "Role",
        "Organization",
        "Type",
        "Interest",
        "Influence",
        "Power-Interest Quadrant",
        "Current Engagement",
        "Desired Engagement",
        "Expectations",
        "Responsibilities",
        "Communication Needs",
        "Communication Frequency",
        "Communication Channel",
        "Owner",
        "Source Reference",
        "Status",
        "Notes",
    ]

    register_sheet.append(register_headers)
    _apply_header_style(
        register_sheet,
        1,
    )

    stakeholders = content.get(
        "stakeholders",
        [],
    )

    for stakeholder in stakeholders:
        register_sheet.append(
            [
                stakeholder.get(
                    "stakeholder_id",
                    "",
                ),
                stakeholder.get("name", ""),
                stakeholder.get("role", ""),
                stakeholder.get(
                    "organization",
                    "",
                ),
                stakeholder.get(
                    "stakeholder_type",
                    "",
                ),
                stakeholder.get(
                    "interest_level",
                    "",
                ),
                stakeholder.get(
                    "influence_level",
                    "",
                ),
                stakeholder.get(
                    "power_interest_quadrant",
                    "",
                ),
                stakeholder.get(
                    "current_engagement",
                    "",
                ),
                stakeholder.get(
                    "desired_engagement",
                    "",
                ),
                _list_to_text(
                    stakeholder.get(
                        "expectations",
                        [],
                    )
                ),
                _list_to_text(
                    stakeholder.get(
                        "responsibilities",
                        [],
                    )
                ),
                _list_to_text(
                    stakeholder.get(
                        "communication_needs",
                        [],
                    )
                ),
                stakeholder.get(
                    "communication_frequency",
                    "",
                ),
                stakeholder.get(
                    "communication_channel",
                    "",
                ),
                stakeholder.get("owner", ""),
                stakeholder.get(
                    "source_reference",
                    "",
                ),
                stakeholder.get("status", ""),
                stakeholder.get("notes", ""),
            ]
        )

    register_sheet.freeze_panes = "A2"
    register_sheet.auto_filter.ref = (
        register_sheet.dimensions
    )

    for row in register_sheet.iter_rows(
        min_row=2
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    _set_column_widths(
        register_sheet,
        {
            "A": 16,
            "B": 24,
            "C": 34,
            "D": 22,
            "E": 16,
            "F": 12,
            "G": 12,
            "H": 22,
            "I": 18,
            "J": 18,
            "K": 42,
            "L": 44,
            "M": 42,
            "N": 22,
            "O": 34,
            "P": 22,
            "Q": 40,
            "R": 14,
            "S": 42,
        },
    )

    summary_sheet.append(
        ["Stakeholder Register Summary", ""]
    )
    summary_sheet.merge_cells(
        "A1:B1"
    )

    summary_sheet["A1"].font = Font(
        bold=True,
        size=14,
    )

    summary_sheet.append(
        ["Project Title", content.get(
            "project_title",
            "",
        )]
    )
    summary_sheet.append(
        ["Purpose", content.get(
            "register_purpose",
            "",
        )]
    )
    summary_sheet.append(
        [
            "Total Stakeholders",
            content.get(
                "total_stakeholders",
                0,
            ),
        ]
    )
    summary_sheet.append(
        [
            "High Influence Stakeholders",
            content.get(
                "high_influence_count",
                0,
            ),
        ]
    )
    summary_sheet.append(
        [
            "Manage Closely Stakeholders",
            content.get(
                "manage_closely_count",
                0,
            ),
        ]
    )
    summary_sheet.append(
        [
            "Artifact Status",
            content.get(
                "artifact_status",
                "Draft",
            ),
        ]
    )

    for row in summary_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    summary_sheet["A2"].font = Font(
        bold=True
    )
    summary_sheet["A3"].font = Font(
        bold=True
    )
    summary_sheet["A4"].font = Font(
        bold=True
    )
    summary_sheet["A5"].font = Font(
        bold=True
    )
    summary_sheet["A6"].font = Font(
        bold=True
    )
    summary_sheet["A7"].font = Font(
        bold=True
    )

    _set_column_widths(
        summary_sheet,
        {
            "A": 34,
            "B": 90,
        },
    )

    assumptions_sheet.append(
        ["Assumptions"]
    )
    _apply_header_style(
        assumptions_sheet,
        1,
    )

    for assumption in content.get(
        "assumptions",
        [],
    ):
        assumptions_sheet.append(
            [assumption]
        )

    next_row = (
        assumptions_sheet.max_row + 2
    )

    assumptions_sheet.cell(
        row=next_row,
        column=1,
        value="Notes",
    )

    assumptions_sheet.cell(
        row=next_row,
        column=1,
    ).font = Font(
        bold=True,
        color="FFFFFF",
    )

    assumptions_sheet.cell(
        row=next_row,
        column=1,
    ).fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    for note in content.get(
        "notes",
        [],
    ):
        assumptions_sheet.append(
            [note]
        )

    for row in assumptions_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    assumptions_sheet.column_dimensions[
        get_column_letter(1)
    ].width = 110

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output