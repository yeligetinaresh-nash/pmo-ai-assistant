from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def _list_to_text(values: list[str]) -> str:
    if not values:
        return ""

    return "\n".join(
        f"• {value}"
        for value in values
    )


def _apply_header_style(
    worksheet,
    row_number: int,
) -> None:
    for cell in worksheet[row_number]:
        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )


def _set_column_widths(
    worksheet,
    widths: dict[str, float],
) -> None:
    for column_letter, width in widths.items():
        worksheet.column_dimensions[
            column_letter
        ].width = width


def generate_raci_matrix_excel(
    content: dict,
) -> BytesIO:
    workbook = Workbook()

    activity_sheet = workbook.active
    activity_sheet.title = "RACI Activities"

    matrix_sheet = workbook.create_sheet(
        "RACI Matrix"
    )

    summary_sheet = workbook.create_sheet(
        "Summary"
    )

    assumptions_sheet = workbook.create_sheet(
        "Assumptions and Notes"
    )

    activity_headers = [
        "Activity ID",
        "Activity Name",
        "Phase",
        "Deliverable / Outcome",
        "Responsible",
        "Accountable",
        "Consulted",
        "Informed",
        "Source Reference",
        "Status",
        "Notes",
    ]

    activity_sheet.append(activity_headers)

    _apply_header_style(
        activity_sheet,
        1,
    )

    activities = content.get(
        "activities",
        [],
    )

    for activity in activities:
        activity_sheet.append(
            [
                activity.get(
                    "activity_id",
                    "",
                ),
                activity.get(
                    "activity_name",
                    "",
                ),
                activity.get(
                    "phase",
                    "",
                ),
                activity.get(
                    "deliverable_or_outcome",
                    "",
                ),
                _list_to_text(
                    activity.get(
                        "responsible",
                        [],
                    )
                ),
                _list_to_text(
                    activity.get(
                        "accountable",
                        [],
                    )
                ),
                _list_to_text(
                    activity.get(
                        "consulted",
                        [],
                    )
                ),
                _list_to_text(
                    activity.get(
                        "informed",
                        [],
                    )
                ),
                activity.get(
                    "source_reference",
                    "",
                ),
                activity.get(
                    "status",
                    "",
                ),
                activity.get(
                    "notes",
                    "",
                ),
            ]
        )

    activity_sheet.freeze_panes = "A2"
    activity_sheet.auto_filter.ref = (
        activity_sheet.dimensions
    )

    for row in activity_sheet.iter_rows(
        min_row=2
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    _set_column_widths(
        activity_sheet,
        {
            "A": 14,
            "B": 42,
            "C": 16,
            "D": 48,
            "E": 24,
            "F": 24,
            "G": 28,
            "H": 28,
            "I": 38,
            "J": 16,
            "K": 48,
        },
    )

    stakeholders = content.get(
        "stakeholders",
        [],
    )

    matrix_headers = [
        "Activity ID",
        "Activity Name",
    ] + stakeholders

    matrix_sheet.append(
        matrix_headers
    )

    _apply_header_style(
        matrix_sheet,
        1,
    )

    for activity in activities:
        row = [
            activity.get(
                "activity_id",
                "",
            ),
            activity.get(
                "activity_name",
                "",
            ),
        ]

        for stakeholder in stakeholders:
            roles = []

            if stakeholder in activity.get(
                "responsible",
                [],
            ):
                roles.append("R")

            if stakeholder in activity.get(
                "accountable",
                [],
            ):
                roles.append("A")

            if stakeholder in activity.get(
                "consulted",
                [],
            ):
                roles.append("C")

            if stakeholder in activity.get(
                "informed",
                [],
            ):
                roles.append("I")

            row.append(
                "/".join(roles)
            )

        matrix_sheet.append(row)

    matrix_sheet.freeze_panes = "C2"
    matrix_sheet.auto_filter.ref = (
        matrix_sheet.dimensions
    )

    for row in matrix_sheet.iter_rows(
        min_row=2
    ):
        for cell in row:
            cell.alignment = Alignment(
                horizontal="center"
                if cell.column >= 3
                else "left",
                vertical="center",
                wrap_text=True,
            )

    matrix_sheet.column_dimensions["A"].width = 14
    matrix_sheet.column_dimensions["B"].width = 48

    for column_number in range(
        3,
        len(stakeholders) + 3,
    ):
        column_letter = (
            matrix_sheet.cell(
                row=1,
                column=column_number,
            ).column_letter
        )

        matrix_sheet.column_dimensions[
            column_letter
        ].width = 26

    summary_sheet.append(
        ["RACI Matrix Summary", ""]
    )

    summary_sheet.merge_cells(
        "A1:B1"
    )

    summary_sheet["A1"].font = Font(
        bold=True,
        size=14,
    )

    summary_sheet.append(
        [
            "Project Title",
            content.get(
                "project_title",
                "",
            ),
        ]
    )

    summary_sheet.append(
        [
            "Purpose",
            content.get(
                "matrix_purpose",
                "",
            ),
        ]
    )

    summary_sheet.append(
        [
            "Total Activities",
            content.get(
                "total_activities",
                0,
            ),
        ]
    )

    summary_sheet.append(
        [
            "Total Stakeholders",
            len(stakeholders),
        ]
    )

    summary_sheet.append(
        [
            "Artifact Status",
            content.get(
                "artifact_status",
                "Draft",
            ),
        ]
    )

    for row_number in range(
        2,
        7,
    ):
        summary_sheet.cell(
            row=row_number,
            column=1,
        ).font = Font(
            bold=True
        )

    for row in summary_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    _set_column_widths(
        summary_sheet,
        {
            "A": 30,
            "B": 90,
        },
    )

    assumptions_sheet.append(
        ["Assumptions"]
    )

    _apply_header_style(
        assumptions_sheet,
        1,
    )

    for assumption in content.get(
        "assumptions",
        [],
    ):
        assumptions_sheet.append(
            [assumption]
        )

    notes_row = (
        assumptions_sheet.max_row + 2
    )

    assumptions_sheet.cell(
        row=notes_row,
        column=1,
        value="Notes",
    )

    assumptions_sheet.cell(
        row=notes_row,
        column=1,
    ).font = Font(
        bold=True,
        color="FFFFFF",
    )

    assumptions_sheet.cell(
        row=notes_row,
        column=1,
    ).fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    for note in content.get(
        "notes",
        [],
    ):
        assumptions_sheet.append(
            [note]
        )

    assumptions_sheet.column_dimensions[
        "A"
    ].width = 110

    for row in assumptions_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return output