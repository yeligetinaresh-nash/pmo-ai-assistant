from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


def _normalize_text(value) -> str:
    if value is None:
        return ""

    return str(value).strip()


def _join_list(values) -> str:
    if not values:
        return ""

    return "\n".join(
        _normalize_text(value)
        for value in values
        if _normalize_text(value)
    )


def generate_requirements_register_excel(
    requirements_register: dict,
    output_path: str | Path,
) -> Path:
    output_path = Path(output_path)

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()

    register_sheet = workbook.active
    register_sheet.title = "Requirements Register"

    headers = [
        "Requirement ID",
        "Requirement Type",
        "Description",
        "Priority",
        "Status",
        "Owner",
        "Source Reference",
        "Test Reference",
        "Acceptance Criteria",
        "Dependencies",
        "Notes",
    ]

    register_sheet.append(headers)

    items = requirements_register.get(
        "items",
        [],
    )

    for item in items:
        register_sheet.append(
            [
                _normalize_text(
                    item.get("requirement_id")
                ),
                _normalize_text(
                    item.get("requirement_type")
                ),
                _normalize_text(
                    item.get("description")
                ),
                _normalize_text(
                    item.get("priority")
                ),
                _normalize_text(
                    item.get("status")
                ),
                _normalize_text(
                    item.get("owner")
                ),
                _normalize_text(
                    item.get("source_reference")
                ),
                _normalize_text(
                    item.get("test_reference")
                ),
                _join_list(
                    item.get(
                        "acceptance_criteria",
                        [],
                    )
                ),
                _join_list(
                    item.get(
                        "dependencies",
                        [],
                    )
                ),
                _normalize_text(
                    item.get("notes")
                ),
            ]
        )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in register_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    register_sheet.freeze_panes = "A2"

    register_sheet.auto_filter.ref = (
        register_sheet.dimensions
    )

    if items:
        table_reference = (
            f"A1:K{len(items) + 1}"
        )

        table = Table(
            displayName="RequirementsRegisterTable",
            ref=table_reference,
        )

        table_style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        table.tableStyleInfo = table_style

        register_sheet.add_table(table)

    column_widths = {
        "A": 16,
        "B": 20,
        "C": 45,
        "D": 10,
        "E": 20,
        "F": 22,
        "G": 42,
        "H": 18,
        "I": 45,
        "J": 20,
        "K": 50,
    }

    for column, width in column_widths.items():
        register_sheet.column_dimensions[
            column
        ].width = width

    for row in register_sheet.iter_rows(
        min_row=2,
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    register_sheet.row_dimensions[1].height = 32

    summary_sheet = workbook.create_sheet(
        "Summary"
    )

    summary_rows = [
        (
            "Project Title",
            requirements_register.get(
                "project_title",
                "",
            ),
        ),
        (
            "Artifact Status",
            requirements_register.get(
                "artifact_status",
                "Draft",
            ),
        ),
        (
            "Purpose",
            requirements_register.get(
                "register_purpose",
                "",
            ),
        ),
        (
            "Total Requirements",
            requirements_register.get(
                "total_requirements",
                0,
            ),
        ),
        (
            "Complete",
            requirements_register.get(
                "complete_count",
                0,
            ),
        ),
        (
            "Partially Complete",
            requirements_register.get(
                "partially_complete_count",
                0,
            ),
        ),
        (
            "Planned",
            requirements_register.get(
                "planned_count",
                0,
            ),
        ),
    ]

    for label, value in summary_rows:
        summary_sheet.append(
            [
                label,
                value,
            ]
        )

    for row in summary_sheet.iter_rows(
        min_row=1,
        max_col=2,
    ):
        row[0].font = Font(
            bold=True
        )

        row[0].fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )

        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    summary_sheet.column_dimensions[
        "A"
    ].width = 28

    summary_sheet.column_dimensions[
        "B"
    ].width = 85

    assumptions_sheet = workbook.create_sheet(
        "Assumptions and Notes"
    )

    assumptions_sheet.append(
        ["Assumptions"]
    )

    assumptions_sheet["A1"].font = Font(
        bold=True,
        size=14,
    )

    for assumption in requirements_register.get(
        "assumptions",
        [],
    ):
        assumptions_sheet.append(
            [_normalize_text(assumption)]
        )

    notes_start_row = (
        assumptions_sheet.max_row + 2
    )

    assumptions_sheet.cell(
        row=notes_start_row,
        column=1,
        value="Notes",
    )

    assumptions_sheet.cell(
        row=notes_start_row,
        column=1,
    ).font = Font(
        bold=True,
        size=14,
    )

    for note in requirements_register.get(
        "notes",
        [],
    ):
        assumptions_sheet.append(
            [_normalize_text(note)]
        )

    assumptions_sheet.column_dimensions[
        "A"
    ].width = 110

    for row in assumptions_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    workbook.save(
        output_path
    )

    return output_path