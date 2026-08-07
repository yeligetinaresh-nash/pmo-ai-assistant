from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


def _normalize_text(value) -> str:
    if value is None:
        return ""

    return str(value).strip()


def _join_list(values) -> str:
    if not values:
        return ""

    return "\n".join(
        _normalize_text(value)
        for value in values
        if _normalize_text(value)
    )


def _style_header_row(
    worksheet,
) -> None:
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.row_dimensions[1].height = 32


def _apply_body_alignment(
    worksheet,
) -> None:
    for row in worksheet.iter_rows(
        min_row=2,
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )


def generate_raid_risk_register_excel(
    raid_risk_register: dict,
    output_path: str | Path,
) -> Path:
    output_path = Path(output_path)

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()

    raid_sheet = workbook.active
    raid_sheet.title = "RAID Log"

    raid_headers = [
        "Item ID",
        "Category",
        "Title",
        "Description",
        "Status",
        "Priority",
        "Owner",
        "Source Reference",
        "Response / Action",
        "Due Date",
        "Dependencies",
        "Notes",
    ]

    raid_sheet.append(
        raid_headers
    )

    raid_items = raid_risk_register.get(
        "raid_items",
        [],
    )

    for item in raid_items:
        raid_sheet.append(
            [
                _normalize_text(
                    item.get("item_id")
                ),
                _normalize_text(
                    item.get("category")
                ),
                _normalize_text(
                    item.get("title")
                ),
                _normalize_text(
                    item.get("description")
                ),
                _normalize_text(
                    item.get("status")
                ),
                _normalize_text(
                    item.get("priority")
                ),
                _normalize_text(
                    item.get("owner")
                ),
                _normalize_text(
                    item.get("source_reference")
                ),
                _normalize_text(
                    item.get("response_or_action")
                ),
                _normalize_text(
                    item.get("due_date")
                ),
                _join_list(
                    item.get(
                        "dependencies",
                        [],
                    )
                ),
                _normalize_text(
                    item.get("notes")
                ),
            ]
        )

    _style_header_row(
        raid_sheet
    )

    raid_sheet.freeze_panes = "A2"
    raid_sheet.auto_filter.ref = (
        raid_sheet.dimensions
    )

    if raid_items:
        raid_table = Table(
            displayName="RAIDLogTable",
            ref=f"A1:L{len(raid_items) + 1}",
        )

        raid_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        raid_sheet.add_table(
            raid_table
        )

    raid_widths = {
        "A": 12,
        "B": 14,
        "C": 34,
        "D": 45,
        "E": 16,
        "F": 12,
        "G": 22,
        "H": 42,
        "I": 48,
        "J": 14,
        "K": 18,
        "L": 42,
    }

    for column, width in raid_widths.items():
        raid_sheet.column_dimensions[
            column
        ].width = width

    _apply_body_alignment(
        raid_sheet
    )

    risk_sheet = workbook.create_sheet(
        "Risk Register"
    )

    risk_headers = [
        "Risk ID",
        "Risk Title",
        "Risk Description",
        "Category",
        "Probability",
        "Impact",
        "Risk Score",
        "Priority",
        "Status",
        "Owner",
        "Mitigation Plan",
        "Contingency Plan",
        "Trigger",
        "Source Reference",
        "Target Date",
        "Notes",
    ]

    risk_sheet.append(
        risk_headers
    )

    risks = raid_risk_register.get(
        "risk_register",
        [],
    )

    for risk in risks:
        risk_sheet.append(
            [
                _normalize_text(
                    risk.get("risk_id")
                ),
                _normalize_text(
                    risk.get("risk_title")
                ),
                _normalize_text(
                    risk.get("risk_description")
                ),
                _normalize_text(
                    risk.get("category")
                ),
                _normalize_text(
                    risk.get("probability")
                ),
                _normalize_text(
                    risk.get("impact")
                ),
                risk.get(
                    "risk_score",
                    0,
                ),
                _normalize_text(
                    risk.get("priority")
                ),
                _normalize_text(
                    risk.get("status")
                ),
                _normalize_text(
                    risk.get("owner")
                ),
                _normalize_text(
                    risk.get("mitigation_plan")
                ),
                _normalize_text(
                    risk.get("contingency_plan")
                ),
                _normalize_text(
                    risk.get("trigger")
                ),
                _normalize_text(
                    risk.get("source_reference")
                ),
                _normalize_text(
                    risk.get("target_date")
                ),
                _normalize_text(
                    risk.get("notes")
                ),
            ]
        )

    _style_header_row(
        risk_sheet
    )

    risk_sheet.freeze_panes = "A2"
    risk_sheet.auto_filter.ref = (
        risk_sheet.dimensions
    )

    if risks:
        risk_table = Table(
            displayName="RiskRegisterTable",
            ref=f"A1:P{len(risks) + 1}",
        )

        risk_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        risk_sheet.add_table(
            risk_table
        )

    risk_widths = {
        "A": 12,
        "B": 34,
        "C": 46,
        "D": 16,
        "E": 14,
        "F": 12,
        "G": 12,
        "H": 12,
        "I": 16,
        "J": 22,
        "K": 50,
        "L": 50,
        "M": 48,
        "N": 42,
        "O": 14,
        "P": 36,
    }

    for column, width in risk_widths.items():
        risk_sheet.column_dimensions[
            column
        ].width = width

    _apply_body_alignment(
        risk_sheet
    )

    summary_sheet = workbook.create_sheet(
        "Summary"
    )

    summary_rows = [
        (
            "Project Title",
            raid_risk_register.get(
                "project_title",
                "",
            ),
        ),
        (
            "Artifact Status",
            raid_risk_register.get(
                "artifact_status",
                "Draft",
            ),
        ),
        (
            "Purpose",
            raid_risk_register.get(
                "register_purpose",
                "",
            ),
        ),
        (
            "Total RAID Items",
            raid_risk_register.get(
                "total_raid_items",
                0,
            ),
        ),
        (
            "Total Risks",
            raid_risk_register.get(
                "total_risks",
                0,
            ),
        ),
        (
            "Open Risks",
            raid_risk_register.get(
                "open_risk_count",
                0,
            ),
        ),
        (
            "High/Critical Risks",
            raid_risk_register.get(
                "high_priority_risk_count",
                0,
            ),
        ),
    ]

    for label, value in summary_rows:
        summary_sheet.append(
            [
                label,
                value,
            ]
        )

    label_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    for row in summary_sheet.iter_rows(
        min_row=1,
        max_col=2,
    ):
        row[0].font = Font(
            bold=True
        )

        row[0].fill = label_fill

        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    summary_sheet.column_dimensions[
        "A"
    ].width = 28

    summary_sheet.column_dimensions[
        "B"
    ].width = 90

    assumptions_sheet = workbook.create_sheet(
        "Assumptions and Notes"
    )

    assumptions_sheet.append(
        ["Assumptions"]
    )

    assumptions_sheet["A1"].font = Font(
        bold=True,
        size=14,
    )

    for assumption in raid_risk_register.get(
        "assumptions",
        [],
    ):
        assumptions_sheet.append(
            [
                _normalize_text(
                    assumption
                )
            ]
        )

    notes_start_row = (
        assumptions_sheet.max_row + 2
    )

    assumptions_sheet.cell(
        row=notes_start_row,
        column=1,
        value="Notes",
    )

    assumptions_sheet.cell(
        row=notes_start_row,
        column=1,
    ).font = Font(
        bold=True,
        size=14,
    )

    for note in raid_risk_register.get(
        "notes",
        [],
    ):
        assumptions_sheet.append(
            [
                _normalize_text(
                    note
                )
            ]
        )

    assumptions_sheet.column_dimensions[
        "A"
    ].width = 110

    for row in assumptions_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    workbook.save(
        output_path
    )

    return output_path