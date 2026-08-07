from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _list_to_text(values: list[str]) -> str:
    if not values:
        return ""

    return ", ".join(values)


def _apply_header_style(
    worksheet,
    row_number: int,
) -> None:
    for cell in worksheet[row_number]:
        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )


def _set_column_widths(
    worksheet,
    widths: dict[str, float],
) -> None:
    for column_letter, width in widths.items():
        worksheet.column_dimensions[
            column_letter
        ].width = width


def generate_project_timeline_excel(
    content: dict,
) -> BytesIO:
    workbook = Workbook()

    timeline_sheet = workbook.active
    timeline_sheet.title = "Project Timeline"

    gantt_sheet = workbook.create_sheet(
        "Gantt View"
    )

    milestones_sheet = workbook.create_sheet(
        "Milestones"
    )

    summary_sheet = workbook.create_sheet(
        "Summary"
    )

    notes_sheet = workbook.create_sheet(
        "Assumptions and Notes"
    )

    activities = content.get(
        "activities",
        [],
    )

    timeline_headers = [
        "Activity ID",
        "Activity Name",
        "Phase",
        "Start Week",
        "End Week",
        "Duration Weeks",
        "Predecessors",
        "Owner",
        "Status",
        "Progress %",
        "Milestone",
        "Milestone Name",
        "Deliverable / Outcome",
        "Source Reference",
        "Notes",
    ]

    timeline_sheet.append(
        timeline_headers
    )

    _apply_header_style(
        timeline_sheet,
        1,
    )

    for activity in activities:
        timeline_sheet.append(
            [
                activity.get("activity_id", ""),
                activity.get("activity_name", ""),
                activity.get("phase", ""),
                activity.get("start_week", 1),
                activity.get("end_week", 1),
                activity.get("duration_weeks", 1),
                _list_to_text(
                    activity.get(
                        "predecessor_ids",
                        [],
                    )
                ),
                activity.get("owner", ""),
                activity.get("status", ""),
                activity.get(
                    "progress_percent",
                    0,
                ),
                "Yes"
                if activity.get("milestone")
                else "No",
                activity.get(
                    "milestone_name",
                    "",
                ),
                activity.get(
                    "deliverable_or_outcome",
                    "",
                ),
                activity.get(
                    "source_reference",
                    "",
                ),
                activity.get("notes", ""),
            ]
        )

    timeline_sheet.freeze_panes = "A2"
    timeline_sheet.auto_filter.ref = (
        timeline_sheet.dimensions
    )

    for row in timeline_sheet.iter_rows(
        min_row=2
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    _set_column_widths(
        timeline_sheet,
        {
            "A": 14,
            "B": 42,
            "C": 16,
            "D": 12,
            "E": 12,
            "F": 14,
            "G": 20,
            "H": 22,
            "I": 16,
            "J": 12,
            "K": 12,
            "L": 34,
            "M": 52,
            "N": 38,
            "O": 48,
        },
    )

    total_duration_weeks = content.get(
        "total_duration_weeks",
        1,
    )

    gantt_headers = [
        "Activity ID",
        "Activity Name",
        "Status",
        "Progress %",
    ]

    gantt_headers.extend(
        [
            f"W{week}"
            for week in range(
                1,
                total_duration_weeks + 1,
            )
        ]
    )

    gantt_sheet.append(
        gantt_headers
    )

    _apply_header_style(
        gantt_sheet,
        1,
    )

    complete_fill = PatternFill(
        fill_type="solid",
        fgColor="70AD47",
    )

    in_progress_fill = PatternFill(
        fill_type="solid",
        fgColor="FFC000",
    )

    planned_fill = PatternFill(
        fill_type="solid",
        fgColor="5B9BD5",
    )

    milestone_fill = PatternFill(
        fill_type="solid",
        fgColor="7030A0",
    )

    for activity in activities:
        row_data = [
            activity.get("activity_id", ""),
            activity.get("activity_name", ""),
            activity.get("status", ""),
            activity.get(
                "progress_percent",
                0,
            ),
        ]

        row_data.extend(
            [""]
            * total_duration_weeks
        )

        gantt_sheet.append(
            row_data
        )

        current_row = gantt_sheet.max_row

        start_week = activity.get(
            "start_week",
            1,
        )

        end_week = activity.get(
            "end_week",
            start_week,
        )

        status = activity.get(
            "status",
            "Planned",
        )

        for week in range(
            start_week,
            end_week + 1,
        ):
            cell = gantt_sheet.cell(
                row=current_row,
                column=4 + week,
            )

            cell.value = "◆" if (
                activity.get("milestone")
                and week == end_week
            ) else "■"

            if (
                activity.get("milestone")
                and week == end_week
            ):
                cell.fill = milestone_fill
            elif status == "Complete":
                cell.fill = complete_fill
            elif status == "In Progress":
                cell.fill = in_progress_fill
            else:
                cell.fill = planned_fill

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

            cell.font = Font(
                color="FFFFFF",
                bold=True,
            )

    gantt_sheet.freeze_panes = "E2"
    gantt_sheet.auto_filter.ref = (
        gantt_sheet.dimensions
    )

    gantt_sheet.column_dimensions["A"].width = 14
    gantt_sheet.column_dimensions["B"].width = 42
    gantt_sheet.column_dimensions["C"].width = 16
    gantt_sheet.column_dimensions["D"].width = 12

    for column_number in range(
        5,
        total_duration_weeks + 5,
    ):
        column_letter = get_column_letter(
            column_number
        )

        gantt_sheet.column_dimensions[
            column_letter
        ].width = 5

    milestone_headers = [
        "Activity ID",
        "Milestone Name",
        "Target Week",
        "Status",
        "Progress %",
        "Owner",
        "Deliverable / Outcome",
    ]

    milestones_sheet.append(
        milestone_headers
    )

    _apply_header_style(
        milestones_sheet,
        1,
    )

    for activity in activities:
        if not activity.get("milestone"):
            continue

        milestones_sheet.append(
            [
                activity.get(
                    "activity_id",
                    "",
                ),
                activity.get(
                    "milestone_name",
                    "",
                ),
                activity.get(
                    "end_week",
                    1,
                ),
                activity.get(
                    "status",
                    "",
                ),
                activity.get(
                    "progress_percent",
                    0,
                ),
                activity.get(
                    "owner",
                    "",
                ),
                activity.get(
                    "deliverable_or_outcome",
                    "",
                ),
            ]
        )

    milestones_sheet.freeze_panes = "A2"
    milestones_sheet.auto_filter.ref = (
        milestones_sheet.dimensions
    )

    for row in milestones_sheet.iter_rows(
        min_row=2
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    _set_column_widths(
        milestones_sheet,
        {
            "A": 14,
            "B": 40,
            "C": 14,
            "D": 16,
            "E": 12,
            "F": 22,
            "G": 60,
        },
    )

    summary_sheet.append(
        ["Project Timeline Summary", ""]
    )

    summary_sheet.merge_cells(
        "A1:B1"
    )

    summary_sheet["A1"].font = Font(
        bold=True,
        size=14,
    )

    summary_rows = [
        (
            "Project Title",
            content.get(
                "project_title",
                "",
            ),
        ),
        (
            "Purpose",
            content.get(
                "timeline_purpose",
                "",
            ),
        ),
        (
            "Planning Basis",
            content.get(
                "planning_basis",
                "",
            ),
        ),
        (
            "Total Duration (Weeks)",
            content.get(
                "total_duration_weeks",
                0,
            ),
        ),
        (
            "Total Activities",
            content.get(
                "total_activities",
                0,
            ),
        ),
        (
            "Total Milestones",
            content.get(
                "total_milestones",
                0,
            ),
        ),
        (
            "Artifact Status",
            content.get(
                "artifact_status",
                "Draft",
            ),
        ),
    ]

    for label, value in summary_rows:
        summary_sheet.append(
            [label, value]
        )

    for row_number in range(
        2,
        summary_sheet.max_row + 1,
    ):
        summary_sheet.cell(
            row=row_number,
            column=1,
        ).font = Font(
            bold=True
        )

    for row in summary_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    _set_column_widths(
        summary_sheet,
        {
            "A": 32,
            "B": 100,
        },
    )

    notes_sheet.append(
        ["Assumptions"]
    )

    _apply_header_style(
        notes_sheet,
        1,
    )

    for assumption in content.get(
        "assumptions",
        [],
    ):
        notes_sheet.append(
            [assumption]
        )

    notes_row = (
        notes_sheet.max_row + 2
    )

    notes_sheet.cell(
        row=notes_row,
        column=1,
        value="Notes",
    )

    notes_sheet.cell(
        row=notes_row,
        column=1,
    ).font = Font(
        bold=True,
        color="FFFFFF",
    )

    notes_sheet.cell(
        row=notes_row,
        column=1,
    ).fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    for note in content.get(
        "notes",
        [],
    ):
        notes_sheet.append(
            [note]
        )

    notes_sheet.column_dimensions[
        "A"
    ].width = 115

    for row in notes_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return output