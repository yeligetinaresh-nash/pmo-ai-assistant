from openpyxl import load_workbook

from app.exporters.project_timeline_excel import (
    generate_project_timeline_excel,
)
from app.exporters.raci_matrix_excel import (
    generate_raci_matrix_excel,
)
from app.exporters.stakeholder_register_excel import (
    generate_stakeholder_register_excel,
)


def test_stakeholder_register_excel_export() -> None:
    content = {
        "project_title": "Test Project",
        "register_purpose": "Test purpose",
        "total_stakeholders": 1,
        "high_influence_count": 1,
        "manage_closely_count": 1,
        "stakeholders": [
            {
                "stakeholder_id": "STK-001",
                "name": "Naresh Yeligeti",
                "role": "Project Manager",
                "organization": "",
                "stakeholder_type": "Internal",
                "interest_level": "High",
                "influence_level": "High",
                "power_interest_quadrant": "Manage Closely",
                "current_engagement": "Leading",
                "desired_engagement": "Leading",
                "expectations": ["Deliver MVP"],
                "responsibilities": ["Manage delivery"],
                "communication_needs": ["Weekly status"],
                "communication_frequency": "Weekly",
                "communication_channel": "Email",
                "owner": "Naresh Yeligeti",
                "source_reference": "BRD",
                "status": "Active",
                "notes": "",
            }
        ],
        "assumptions": [],
        "notes": [],
        "artifact_status": "Draft",
    }

    output = generate_stakeholder_register_excel(content)

    workbook = load_workbook(output)

    assert "Stakeholder Register" in workbook.sheetnames
    assert "Summary" in workbook.sheetnames
    assert workbook["Stakeholder Register"]["A2"].value == "STK-001"


def test_raci_matrix_excel_export() -> None:
    content = {
        "project_title": "Test Project",
        "matrix_purpose": "Test purpose",
        "total_activities": 1,
        "stakeholders": ["Naresh Yeligeti"],
        "activities": [
            {
                "activity_id": "RAC-001",
                "activity_name": "Build API",
                "phase": "Development",
                "deliverable_or_outcome": "Working API",
                "responsible": ["Naresh Yeligeti"],
                "accountable": ["Naresh Yeligeti"],
                "consulted": [],
                "informed": [],
                "source_reference": "BRD",
                "status": "Complete",
                "notes": "",
            }
        ],
        "assumptions": [],
        "notes": [],
        "artifact_status": "Draft",
    }

    output = generate_raci_matrix_excel(content)

    workbook = load_workbook(output)

    assert "RACI Activities" in workbook.sheetnames
    assert "RACI Matrix" in workbook.sheetnames
    assert workbook["RACI Matrix"]["C2"].value == "R/A"


def test_project_timeline_excel_export() -> None:
    content = {
        "project_title": "Test Project",
        "timeline_purpose": "Test purpose",
        "planning_basis": "Relative weeks",
        "total_duration_weeks": 2,
        "total_activities": 1,
        "total_milestones": 1,
        "activities": [
            {
                "activity_id": "TL-001",
                "activity_name": "Project setup",
                "phase": "Initiation",
                "start_week": 1,
                "end_week": 2,
                "duration_weeks": 2,
                "predecessor_ids": [],
                "owner": "Naresh Yeligeti",
                "status": "Complete",
                "progress_percent": 100,
                "milestone": True,
                "milestone_name": "Setup Complete",
                "deliverable_or_outcome": "Environment ready",
                "source_reference": "WBS 1",
                "notes": "",
            }
        ],
        "assumptions": [],
        "notes": [],
        "artifact_status": "Draft",
    }

    output = generate_project_timeline_excel(content)

    workbook = load_workbook(output)

    assert "Project Timeline" in workbook.sheetnames
    assert "Gantt View" in workbook.sheetnames
    assert "Milestones" in workbook.sheetnames
    assert workbook["Project Timeline"]["A2"].value == "TL-001"
    assert workbook["Gantt View"]["E2"].value == "■"
    assert workbook["Gantt View"]["F2"].value == "◆"